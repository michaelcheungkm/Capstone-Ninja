import os

from openpyxl import load_workbook
import numpy as np

wb = load_workbook('C:\\Users\\user\\PycharmProjects\\XLparser_capstone\\capstone_question.xlsx')

# final vars
SEPARATOR = ';'  # Separating value in split function for question options
TITLE_IND = 0  # Column index of Titles
QUESTIONS_IND = 1  # Column index of Questions
QUESTION_TYPE_IND = 2  # Column index of Question Types
ANSWERS_IND = 3  # Column index of Options
CORRECT_ANSWER_IND = 4  # Column index of Answers
CORRECT_ANSWER_MCQ_IND = 5  # Column index of MCQ Answers
ANSWER_DESCRIPTION_IND = 6  # Column index of Answer Descriptions
fields = wb.worksheets[0].max_column  # Number of column fields in workbook

MCQ_IDENTIFIER = 'Multiple Choice'  # MCQ type identifier
TF_IDENTIFIER = 'True or false'  # True or false type identifier
TF_ANSWERS = 'True' + SEPARATOR + 'False'  # Answer options for true or false
ASCII_a = 97  # ASCII value for 'a'

PATH_TO_HTML_HEAD = 'template/html/head.txt'  # Path to header html file
PATH_TO_HTML_FOOT = 'template/html/foot.txt'  # Path to footer html file
PATH_TO_JS_HEAD = 'template/js/head.txt'  # Path to header js file
PATH_TO_JS_FOOT = 'template/js/foot.txt'  # Path to footer js file


# Returns 1D array consisting of number of rows in each sheet in a workbook
# twb: Workbook in question
def arrOfRowNums(twb):
    numSheets = len(twb.worksheets)
    arrOut = []
    for x in range(0, numSheets):
        twbsheet = twb.worksheets[x]
        arrOut.append(twbsheet.max_row)
    return arrOut


# Returns value for largest number of rows out of all sheets in a workbook
# arrOfRows: Array of number of rows
def getMaxRows(arrOfRows):
    counter = 0
    for x in arrOfRows:
        if x > counter:
            counter = x
    return counter


# Returns 3D array of entire workbook
# twb: Workbook in question
def genFullArr(twb):
    rowsInSheets = arrOfRowNums(twb)
    sheetMaxRows = getMaxRows(arrOfRowNums(twb))
    sheetArr = np.empty((len(twb.sheetnames) - 1, sheetMaxRows, fields), dtype=object)
    for x in range(0, len(twb.sheetnames) - 1):
        tsheet = twb.worksheets[x]
        for i in range(0, rowsInSheets[x] - 1):
            for j in range(0, fields):
                sheetArr[x, i, j] = tsheet.cell(2 + i, j + 1).value
    return sheetArr


# Returns 2D array of entire sheet
# twb: Workbook containing sheet
# sheetNum: Sheet index within workbook
def genSheetArr(twb, sheetNum):
    rowsInSheets = arrOfRowNums(twb)
    sheetArr = np.empty((rowsInSheets[sheetNum] - 1, fields), dtype=object)
    tsheet = twb.worksheets[sheetNum]
    for i in range(0, rowsInSheets[sheetNum] - 1):
        for j in range(0, fields):
            if tsheet.cell(2 + i, j + 1).value is None:
                sheetArr[i, j] = "--"
            else:
                sheetArr[i, j] = tsheet.cell(2 + i, j + 1).value
        sheetArr[i, 0] = sheetArr[i, 0].replace(' ', '_').replace('-', '_')
    return sheetArr


# Returns 1D array of title column
# twb: Workbook
# sheetNum: Sheet index within workbook
def pullTitleCol(twb, sheetNum):
    sheetArr = genSheetArr(twb, sheetNum)
    numRows = len(sheetArr)
    titleArr = np.empty(numRows, dtype=object)
    for x in range(0, numRows):
        titleArr[x] = sheetArr[x][TITLE_IND]
    return titleArr


# Returns 1D array of unique keywords from titles
# twb: Workbook containing sheet
# sheetNum: Sheet index within workbook
def pullKeywords(twb, sheetNum):
    titleCol = pullTitleCol(twb, sheetNum)
    keywordList = []
    for x in titleCol:
        for i in range(0, 10):
            x = x.replace('_Q' + str(i), '')
        if x not in keywordList:
            keywordList.append(x)
    return keywordList


# Returns 2D array of all rows containing keyword in title
# twb: Workbook containing sheet
# sheetNum: Sheet index within workbook
# keyword: Keyword in question
def pullKeywordQsArr(twb, sheetNum, keyword):
    sheetArr = genSheetArr(twb, sheetNum)
    titleCol = pullTitleCol(twb, sheetNum)
    arrOfKeywordIndex = []
    for x in range(0, len(titleCol)):
        if keyword in titleCol[x]:
            arrOfKeywordIndex.append(x)
    keywordArr = np.empty((len(arrOfKeywordIndex), fields), dtype=object)
    for x in range(0, len(arrOfKeywordIndex)):
        keywordArr[x] = sheetArr[arrOfKeywordIndex[x]]
    return keywordArr


# Sorts array to numerical order
# keywordArr: Array of rows containing keywords in title
def arraySort(keywordArr):
    sortedArr = np.empty((len(keywordArr), fields), dtype=object)
    for x in range(0, len(keywordArr)):
        for i in keywordArr[x][TITLE_IND]:
            for j in range(1, len(keywordArr) + 1):
                if str(j) in i:
                    sortedArr[j - 1] = keywordArr[x]
    return sortedArr


# Populates empty arrays from sorted array
# sortedArr: Sorted array
# QUESTIONS: Empty array to be populated
# ANSWERS: Empty array to be populated
# CORRECT_ANSWER: Empty array to be populated
# ANSWER_DESCRIPTION: Empty array to be populated
def populateArrs(sortedArr, QUESTIONS, ANSWERS, CORRECT_ANSWER, ANSWER_DESCRIPTION):
    for x in range(0, len(sortedArr)):
        QUESTIONS[x] = str(sortedArr[x][QUESTIONS_IND])
        if MCQ_IDENTIFIER in sortedArr[x][QUESTION_TYPE_IND]:
            ANSWERS[x] = str(sortedArr[x][ANSWERS_IND])
            CORRECT_ANSWER[x] = str(sortedArr[x][CORRECT_ANSWER_IND])
        elif TF_IDENTIFIER in sortedArr[x][QUESTION_TYPE_IND]:
            ANSWERS[x] = TF_ANSWERS
            CORRECT_ANSWER[x] = str(sortedArr[x][CORRECT_ANSWER_MCQ_IND])
        ANSWER_DESCRIPTION[x] = str(sortedArr[x][ANSWER_DESCRIPTION_IND])
    for i in range(0, len(CORRECT_ANSWER)):
        if CORRECT_ANSWER[i] in 'True':
            CORRECT_ANSWER[i] = 'a'
        if CORRECT_ANSWER[i] in 'False':
            CORRECT_ANSWER[i] = 'b'
        if '.0' in CORRECT_ANSWER[i]:
            CORRECT_ANSWER[i] = CORRECT_ANSWER[i].replace('.0', '')
            CORRECT_ANSWER[i] = convertNumToCharStr(CORRECT_ANSWER[i])
        for j in range (0, 10):
            if CORRECT_ANSWER[i] == str(j):
                CORRECT_ANSWER[i] = convertNumToCharStr(CORRECT_ANSWER[i])
    for i in range(0, len(ANSWER_DESCRIPTION)):
        if ANSWER_DESCRIPTION[i] in '--':
            ANSWER_DESCRIPTION[i] = ANSWER_DESCRIPTION[i].replace('--', '')


# Turns number in string form into character mapped to lower cased alphabet
# numAsStr: Number in string form
def convertNumToCharStr(numAsStr):
    return str(chr(int(numAsStr) - 1 + ASCII_a))


# Writes arrays to JS code
# KEYWORD: Keyword of arrays
# QUESTIONS: Questions array
# ANSWERS: Answers array
# CORRECT_ANSWER: Correct answers array
# ANSWER_DESCRIPTION: Answer descriptions array
def buildJSArr(KEYWORD, QUESTIONS, ANSWERS, CORRECT_ANSWER, ANSWER_DESCRIPTION):
    JS_QUESTION_ARRAY = ''
    for x in range(0, len(QUESTIONS)):
        JS_QUESTION_ARRAY += '\nconst ' + KEYWORD + '_' + str(x + 1) + ' = [' + '{question: "' + QUESTIONS[
            x] + '", answers: {'
        answersExploded = ANSWERS[x].split(SEPARATOR)
        for i in range(0, len(answersExploded)):
            answersExploded[i] = answersExploded[i].replace('\n', '')
        for i in range(1, len(answersExploded) + 1):
            JS_QUESTION_ARRAY += convertNumToCharStr(str(i)) + ': "' + answersExploded[i - 1] + '",'
        JS_QUESTION_ARRAY += '}, correctAnswer: "' + CORRECT_ANSWER[x] + '"}]'
    JS_QUESTION_ARRAY += '\nlet ' + 'quizArr = ['
    for x in range(0, len(QUESTIONS)):
        JS_QUESTION_ARRAY += KEYWORD + '_' + str(x + 1) + ','
    JS_QUESTION_ARRAY += ']'
    return JS_QUESTION_ARRAY


# Returns JS code to be put into file
# keywordArr: Unsorted array of rows with a keyword
def processKeywordQsArr(keywordArr):
    sortedArr = arraySort(keywordArr)
    numOfQuestions = len(sortedArr)
    KEYWORD = sortedArr[0][0][:-3]
    QUESTIONS = np.empty(numOfQuestions, dtype=object)
    ANSWERS = np.empty(numOfQuestions, dtype=object)
    CORRECT_ANSWER = np.empty(numOfQuestions, dtype=object)
    ANSWER_DESCRIPTION = np.empty(numOfQuestions, dtype=object)
    populateArrs(sortedArr, QUESTIONS, ANSWERS, CORRECT_ANSWER, ANSWER_DESCRIPTION)
    JS_QUESTION_ARRAY = buildJSArr(KEYWORD, QUESTIONS, ANSWERS, CORRECT_ANSWER, ANSWER_DESCRIPTION)
    return JS_QUESTION_ARRAY


# Generates .js files from .xlsx workbook
def generateFiles():
    for x in range(0, len(wb.worksheets)):
        for i in range(0, len(pullKeywords(wb, x))):
            keywordArr = pullKeywordQsArr(wb, x, pullKeywords(wb, x)[i])
            sortedArr = arraySort(keywordArr)
            keyphrase = str(sortedArr[0][0][:-3])
            if not os.path.exists("PARSED"):
                os.makedirs("PARSED")
            if not os.path.exists("PARSED/" + keyphrase):
                os.makedirs("PARSED/" + keyphrase)
            fileNameStr = "PARSED/" + keyphrase + '/' + keyphrase + ".js"
            fileHTMLNameStr = "PARSED/" + keyphrase + '/' + 'quiz' + ".php"
            print('PARSING: ' + fileNameStr.replace("PARSED/", ""))
            jsStr = processKeywordQsArr(pullKeywordQsArr(wb, x, pullKeywords(wb, x)[i]))
            htmlStr = '<script src="' + keyphrase + '.js"></script>'
            f = open(fileNameStr, "w")
            f1 = open(PATH_TO_JS_HEAD, "r")
            f2 = open(PATH_TO_JS_FOOT, "r")
            f3 = open(fileHTMLNameStr, "w")
            f4 = open(PATH_TO_HTML_HEAD, "r")
            f5 = open(PATH_TO_HTML_FOOT, "r")
            f.write(f1.read() + '\n')
            f.write(jsStr)
            f.write('\n\n' + f2.read())
            f3.write(f4.read() + '\n')
            f3.write(htmlStr)
            f3.write('\n\n' + f5.read())
            f.close()
            f1.close()
            f2.close()


generateFiles()
